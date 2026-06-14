import openpyxl
import json
import os
import sys

# Trova i file Excel
dati_dir = './dati'
files = [f for f in os.listdir(dati_dir) if f.endswith('.xlsx')]
print(f"Trovati {len(files)} file Excel: {files}")

data = {}

for f in files:
    path = os.path.join(dati_dir, f)
    wb = openpyxl.load_workbook(path, data_only=True)
    file_data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                row_dict = {}
                for i, h in enumerate(headers):
                    if h and i < len(row):
                        row_dict[str(h)] = row[i]
                rows.append(row_dict)
        file_data[sheet_name] = rows
    data[f] = file_data
    print(f"  {f}: {len(wb.sheetnames)} fogli, {sum(len(v) for v in file_data.values())} righe totali")

# Salva come JSON
with open('dati/imported_data.json', 'w') as f:
    json.dump(data, f, ensure_ascii=False, default=str, indent=2)

print(f"\n✅ Dati salvati in dati/imported_data.json")
